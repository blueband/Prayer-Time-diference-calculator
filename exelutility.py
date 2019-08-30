import openpyxl as exel
import os
dir_src = os.path.dirname(__file__)

def create_workbook():
    # A copy of workbook and one single worksheet is produce to the calling functions

    wkbook = exel.Workbook()
    wsheet = wkbook.active
    return wkbook, wsheet

def xlxsfield_creator(default_town_name,current_sheet, cell_spacer=0):
    # wkbook, current_sheet = create_workbook()
    current_sheet.cell(row=2, column=4+cell_spacer).value = 'Town'
    current_sheet.cell(row=2, column=5+cell_spacer).value = default_town_name
    current_sheet.cell(row=3, column=1 + cell_spacer).value = 'Date'
    current_sheet.cell(row=3, column=2 + cell_spacer).value  = 'Fajr & Shuruq'
    current_sheet.cell(row=3, column=3 + cell_spacer).value = 'Zurh'
    current_sheet.cell(row=3, column=4 + cell_spacer).value = 'Asr'
    current_sheet.cell(row=3, column=5 + cell_spacer).value = 'Magrib'
    current_sheet.cell(row=3, column=6 + cell_spacer).value = 'Isha'


def xlsx_cell(listobj, TOWN_BUCKET, cell_spacer=0):
    wkbook, current_sheet = create_workbook()
    xlsx_row = len(listobj[0][0])
    default_town_data = listobj[0][0]   # Extract ref_data
    default_town_name = listobj[0][1]   # Extract Town_name

    num_row = 0

    while num_row < xlsx_row:
        current_sheet.cell(row=num_row + 4, column=1+cell_spacer).value = default_town_data[num_row][0]
        current_sheet.cell(row=num_row + 4, column=2+cell_spacer).value = default_town_data[num_row][1]
        current_sheet.cell(row=num_row + 4, column=3+cell_spacer).value = default_town_data[num_row][2]
        current_sheet.cell(row=num_row + 4, column=4+cell_spacer).value = default_town_data[num_row][3]
        current_sheet.cell(row=num_row + 4, column=5+cell_spacer).value = default_town_data[num_row][4]
        current_sheet.cell(row=num_row + 4, column=6+cell_spacer).value = default_town_data[num_row][5]
        num_row += 1
    xlxsfield_creator(default_town_name, current_sheet,cell_spacer)
    wkbook.save('data.xlsx')

    ######## Looping to extract value from TOWN_BUCKET here
    cell_spacer = 8
    for town in TOWN_BUCKET[0]:
        data2Cell(current_sheet,town, cell_spacer)
        columHeader(current_sheet, town,cell_spacer)

        cell_spacer += 6
    wkbook.save('data.xlsx')


def columHeader(*args):
    current_sheet = args[0]
    default_town_name = args[1][0][1]
    cell_spacer = args[2]

    current_sheet.cell(row=2, column=4 + cell_spacer).value = 'Town'
    current_sheet.cell(row=2, column=5 + cell_spacer).value = default_town_name
    current_sheet.cell(row=3, column=1 + cell_spacer).value = 'Fajr & Shuruq'
    current_sheet.cell(row=3, column=2 + cell_spacer).value = 'Zurh'
    current_sheet.cell(row=3, column=3 + cell_spacer).value = 'Asr'
    current_sheet.cell(row=3, column=4 + cell_spacer).value = 'Magrib'
    current_sheet.cell(row=3, column=5 + cell_spacer).value = 'Isha'




def data2Cell(*args):
    current_sheet = args[0]
    xlsx_row = len(args[1][0][0])
    default_town_data = args[1][0][0]  # Extract ref_data
    cell_spacer = args[2]

    num_row = 0

    while num_row < xlsx_row:
        current_sheet.cell(row=num_row + 4, column=1 + cell_spacer).value = default_town_data[num_row][0]
        current_sheet.cell(row=num_row + 4, column=2 + cell_spacer).value = default_town_data[num_row][1]
        current_sheet.cell(row=num_row + 4, column=3 + cell_spacer).value = default_town_data[num_row][2]
        current_sheet.cell(row=num_row + 4, column=4 + cell_spacer).value = default_town_data[num_row][3]
        current_sheet.cell(row=num_row + 4, column=5 + cell_spacer).value = default_town_data[num_row][4]
        num_row += 1


# def calc_value_cell(listobj, cell_spacer):
#     wkbook, current_sheet = create_workbook()
#     xlsx_row = len(listobj)
#     num_row = 0
#
#     while num_row < xlsx_row:
#         for value in (listobj[num_row]):
#             current_sheet.cell(row=num_row + 4, column=1 + cell_spacer).value = str(value[0])
#             current_sheet.cell(row=num_row + 4, column=2 + cell_spacer).value = str(value[1])
#             current_sheet.cell(row=num_row + 4, column=3 + cell_spacer).value = str(value[2])
#             current_sheet.cell(row=num_row + 4, column=4 + cell_spacer).value = str(value[3])
#             current_sheet.cell(row=num_row + 4, column=5 + cell_spacer).value = str(value[4])
#             num_row += 1



